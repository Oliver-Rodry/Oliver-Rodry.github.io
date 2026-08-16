#!/usr/bin/env python3
"""Update products.csv from an inventory workbook and produce an audit report."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_COLUMNS = {
    "cbarras",
    "nombre_producto",
    "descripcion_producto",
    "existencia",
    "precio_venta",
    "nombre_categoria",
}
CATALOG_COLUMNS = ["sku", "name", "category", "price_dop", "stock", "description"]


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def number(value: object, label: str, sku: str) -> Decimal:
    try:
        return Decimal(str(value or 0))
    except InvalidOperation as exc:
        raise ValueError(f"Invalid {label} for {sku}: {value!r}") from exc


def display_number(value: Decimal) -> str:
    normalized = value.normalize()
    return str(int(normalized)) if normalized == normalized.to_integral() else format(normalized, "f")


def read_catalog(path: Path) -> tuple[list[dict[str, str]], dict[str, dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if not rows or set(CATALOG_COLUMNS) - set(rows[0]):
        raise ValueError(f"Catalog has missing columns: {path}")
    by_sku = {clean(row["sku"]): row for row in rows}
    if len(by_sku) != len(rows):
        raise ValueError(f"Catalog contains duplicate SKUs: {path}")
    return rows, by_sku


def read_workbook(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(iterator)]
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise ValueError(f"Workbook is missing columns: {', '.join(sorted(missing))}")
    rows = [dict(zip(headers, values)) for values in iterator]
    workbook.close()
    return [row for row in rows if clean(row.get("cbarras"))]


def load_config(path: Path | None) -> dict[str, object]:
    if not path:
        return {"stock_overrides": {}}
    return json.loads(path.read_text(encoding="utf-8"))


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def build_catalog(
    workbook_rows: list[dict[str, object]],
    old_by_sku: dict[str, dict[str, str]],
    overrides: dict[str, dict[str, object]],
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, object]]]:
    output: list[dict[str, str]] = []
    raw_output: list[dict[str, str]] = []
    applied_overrides: list[dict[str, object]] = []
    seen: set[str] = set()
    for source in workbook_rows:
        sku = clean(source["cbarras"])
        if sku in seen:
            raise ValueError(f"Workbook contains duplicate SKU: {sku}")
        seen.add(sku)
        previous = old_by_sku.get(sku, {})
        workbook_stock = number(source["existencia"], "stock", sku)
        stock = workbook_stock
        if sku in overrides:
            forced_stock = number(overrides[sku].get("stock", 0), "override stock", sku)
            applied_overrides.append(
                {
                    "sku": sku,
                    "name": clean(source["nombre_producto"]),
                    "workbook_stock": display_number(stock),
                    "catalog_stock": display_number(forced_stock),
                    "reason": clean(overrides[sku].get("reason")),
                }
            )
            stock = forced_stock
        category = clean(previous.get("category")) or clean(source["nombre_categoria"]) or "CAT. GENERAL"
        row = {
            "sku": sku,
            "name": clean(source["nombre_producto"]),
            "category": category,
            "price_dop": display_number(number(source["precio_venta"], "price", sku)),
            "stock": display_number(stock),
            "description": clean(source["descripcion_producto"]),
        }
        output.append(row)
        raw_output.append({**row, "stock": display_number(workbook_stock)})
    return output, raw_output, applied_overrides


def compare(
    old_by_sku: dict[str, dict[str, str]],
    new_rows: list[dict[str, str]],
    stock_override_skus: set[str],
) -> dict[str, object]:
    new_by_sku = {row["sku"]: row for row in new_rows}
    sold: list[dict[str, object]] = []
    restocked: list[dict[str, object]] = []
    price_changes: list[dict[str, object]] = []
    newly_out: list[dict[str, object]] = []
    new_products = [row for sku, row in new_by_sku.items() if sku not in old_by_sku]
    removed_products = [row for sku, row in old_by_sku.items() if sku not in new_by_sku]
    sold_units = Decimal(0)
    estimated_sales = Decimal(0)

    for sku in sorted(set(old_by_sku) & set(new_by_sku)):
        old, new = old_by_sku[sku], new_by_sku[sku]
        old_stock = number(old["stock"], "old stock", sku)
        new_stock = number(new["stock"], "new stock", sku)
        old_price = number(old["price_dop"], "old price", sku)
        new_price = number(new["price_dop"], "new price", sku)
        change = new_stock - old_stock
        item = {"sku": sku, "name": new["name"]}
        if sku in stock_override_skus:
            change = Decimal(0)
        if change < 0:
            units = -change
            sale_value = units * old_price
            sold.append(
                {
                    **item,
                    "category": new["category"],
                    "units": display_number(units),
                    "unit_price_dop": display_number(old_price),
                    "estimated_dop": display_number(sale_value),
                }
            )
            sold_units += units
            estimated_sales += sale_value
        elif change > 0:
            restocked.append({**item, "units_added": display_number(change)})
        if old_price != new_price:
            price_changes.append({**item, "old_price": display_number(old_price), "new_price": display_number(new_price)})
        if old_stock > 0 and new_stock <= 0:
            newly_out.append(item)

    return {
        "summary": {
            "products_in_catalog": len(new_rows),
            "sold_units_including_services": display_number(sold_units),
            "estimated_sales_dop_including_services": display_number(estimated_sales),
            "sold_product_count": len(sold),
            "restocked_product_count": len(restocked),
            "price_change_count": len(price_changes),
            "newly_out_of_stock_count": len(newly_out),
            "new_product_count": len(new_products),
            "removed_product_count": len(removed_products),
        },
        "sold": sold,
        "restocked": restocked,
        "price_changes": price_changes,
        "newly_out_of_stock": newly_out,
        "new_products": new_products,
        "removed_products": removed_products,
    }


def write_catalog(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CATALOG_COLUMNS, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path, help="Inventory .xlsx file")
    parser.add_argument("--baseline", type=Path, default=Path("products.csv"))
    parser.add_argument("--output", type=Path, default=Path("products.csv"))
    parser.add_argument("--overrides", type=Path, default=Path("inventory_overrides.json"))
    parser.add_argument("--report", type=Path, default=Path("inventory_update_report.json"))
    parser.add_argument("--state", type=Path, help="Optional processed-file state for duplicate protection")
    parser.add_argument("--apply", action="store_true", help="Write catalog and processed state")
    args = parser.parse_args()

    digest = file_sha256(args.input)
    if args.state and args.state.exists():
        state = json.loads(args.state.read_text(encoding="utf-8"))
        if state.get("last_workbook_sha256") == digest:
            print(json.dumps({"status": "duplicate", "sha256": digest}, indent=2))
            return 0

    _, old_by_sku = read_catalog(args.baseline)
    workbook_rows = read_workbook(args.input)
    config = load_config(args.overrides)
    new_rows, raw_rows, applied_overrides = build_catalog(
        workbook_rows, old_by_sku, config.get("stock_overrides", {})
    )
    report = compare(old_by_sku, raw_rows, set(config.get("stock_overrides", {})))
    report.update(
        {
            "status": "applied" if args.apply else "preview",
            "workbook": args.input.name,
            "workbook_sha256": digest,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "manual_overrides": applied_overrides,
        }
    )
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    if args.apply:
        write_catalog(args.output, new_rows)
        if args.state:
            args.state.write_text(
                json.dumps({"last_workbook_sha256": digest, "processed_at": report["generated_at"]}, indent=2) + "\n",
                encoding="utf-8",
            )
    print(json.dumps(report["summary"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (OSError, ValueError, KeyError, json.JSONDecodeError) as error:
        print(f"ERROR: {error}", file=sys.stderr)
        raise SystemExit(1)
